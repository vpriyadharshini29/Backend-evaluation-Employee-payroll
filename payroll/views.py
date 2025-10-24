from django.shortcuts import render, redirect, get_object_or_404
from .models import Employee, Attendance, PayrollRecord
from .forms import EmployeeForm, AttendanceForm
from django.db.models import Count
from django.contrib import messages
from django.http import HttpResponse
import datetime
from decimal import Decimal
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import io
import openpyxl
from openpyxl.utils import get_column_letter

def employee_list(request):
    qs = Employee.objects.all().order_by('first_name')
    return render(request, 'payroll/employee_list.html', {'employees': qs})

def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee created.")
            return redirect('payroll:employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'payroll/employee_form.html', {'form': form, 'title': 'Add Employee'})

def employee_edit(request, pk):
    emp = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=emp)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee updated.")
            return redirect('payroll:employee_list')
    else:
        form = EmployeeForm(instance=emp)
    return render(request, 'payroll/employee_form.html', {'form': form, 'title': 'Edit Employee'})

def attendance_list(request):
    attendances = Attendance.objects.select_related('employee').all()[:200]
    form = AttendanceForm()
    return render(request, 'payroll/attendance_list.html', {'attendances': attendances, 'form': form})

def attendance_mark(request):
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Attendance recorded.")
    return redirect('payroll:attendance_list')

def payroll_list(request):
    payrolls = PayrollRecord.objects.select_related('employee').all()
    return render(request, 'payroll/payroll_list.html', {'payrolls': payrolls})

def generate_payroll(request):
    # naive payroll calculation: net = basic_salary * (days_present / working_days)
    # Expects query params month=YYYY-MM or form post
    month = request.GET.get('month')
    if not month:
        month = datetime.date.today().strftime('%Y-%m')
    year, mon = map(int, month.split('-'))
    first = datetime.date(year, mon, 1)
    last = (first.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)
    working_days = 22  # simple assumption
    created = []
    for emp in Employee.objects.all():
        presents = Attendance.objects.filter(employee=emp, date__range=(first, last), present=True).count()
        gross = emp.basic_salary
        if working_days > 0:
            net = (Decimal(presents) / Decimal(working_days)) * gross
        else:
            net = gross
        rec, _ = PayrollRecord.objects.get_or_create(employee=emp, month=first, defaults={
            'gross_salary': gross, 'net_salary': round(net,2)
        })
        # if exists, update
        if not _:
            rec.gross_salary = gross
            rec.net_salary = round(net,2)
            rec.save()
        created.append(rec)
    messages.success(request, f"Payroll generated for {month}.")
    return redirect('payroll:payroll_list')

def payroll_pdf(request, pk):
    rec = get_object_or_404(PayrollRecord, pk=pk)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setFont('Helvetica-Bold', 16)
    p.drawString(50, 800, 'Salary Slip')
    p.setFont('Helvetica', 12)
    p.drawString(50, 770, f'Employee: {rec.employee}')
    p.drawString(50, 750, f'Month: {rec.month.strftime("%B %Y")}')
    p.drawString(50, 730, f'Gross Salary: {rec.gross_salary}')
    p.drawString(50, 710, f'Net Salary: {rec.net_salary}')
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

def export_payroll_excel(request):
    payrolls = PayrollRecord.objects.select_related('employee').all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"
    headers = ['Employee', 'Month', 'Gross Salary', 'Net Salary']
    ws.append(headers)
    for r in payrolls:
        ws.append([str(r.employee), r.month.strftime('%Y-%m'), float(r.gross_salary), float(r.net_salary)])
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=payroll.xlsx'
    return resp
